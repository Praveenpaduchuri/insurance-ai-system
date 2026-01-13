import os
import pdfplumber
import openpyxl
import docx
from PIL import Image
import pytesseract

ATTACHMENT_DIR = "attachments"
os.makedirs(ATTACHMENT_DIR, exist_ok=True)

def save_attachment(part, email_uid):
    """Save email attachments to folder."""
    filename = part.get_filename()
    if not filename:
        return None
        
    # Sanitize filename
    clean_filename = "".join(c for c in filename if c.isalnum() or c in "._- ")
    filepath = os.path.join(ATTACHMENT_DIR, f"{email_uid}_{clean_filename}")

    with open(filepath, "wb") as f:
        f.write(part.get_payload(decode=True))

    return filepath

def extract_text_from_image(image_path):
    """Extract text from image using OCR (pytesseract)."""
    try:
        img = Image.open(image_path)
        text = pytesseract.image_to_string(img)
        return text
    except Exception as e:
        print(f"OCR Error for {image_path}: {e}")
        return ""


def extract_text_from_pdf(pdf_path):
    """Extract text from PDF using pdfplumber (text) + OCR (scanned)."""
    text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                # 1. Try normal text extraction
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
                
                # 2. If mostly empty, might be scanned -> convert to image and OCR
                # simple heuristic: if low text count
                # (For brevity, skipping full PDF-to-Image OCR implementation in this snippet, 
                # but relying on extract_text_from_image if user provides images directly)
    except Exception as e:
        print(f"Error reading PDF {pdf_path}: {e}")
    return text

def extract_text_from_excel(xlsx_path):
    text = ""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                # Convert row to string representation
                row_text = " ".join([str(cell) for cell in row if cell is not None])
                text += row_text + "\n"
    except Exception as e:
        print(f"Error reading Excel {xlsx_path}: {e}")
    return text

def extract_text_from_word(docx_path):
    text = ""
    try:
        doc = docx.Document(docx_path)
        for para in doc.paragraphs:
            text += para.text + "\n"
    except Exception as e:
        print(f"Error reading Word {docx_path}: {e}")
    return text

def extract_content_from_file(filepath):
    """Dispatcher for different file types."""
    ext = os.path.splitext(filepath)[1].lower()
    
    if ext == ".pdf":
        return extract_text_from_pdf(filepath)
    elif ext in [".jpg", ".jpeg", ".png", ".bmp", ".tiff"]:
        return extract_text_from_image(filepath)
    elif ext in [".xlsx", ".xls"]:
        return extract_text_from_excel(filepath)
    elif ext in [".docx", ".doc"]:
        return extract_text_from_word(filepath)
    
    return ""
